"""
utils/helpers.py
────────────────
Funções auxiliares compartilhadas por todos os módulos.
"""

import os
import re
from pathlib import Path
from datetime import datetime


# ════════════════════════════════════════════════════════
#  VALIDAÇÃO DE ARQUIVO
# ════════════════════════════════════════════════════════
def validar_arquivo(caminho: str) -> tuple[bool, str]:
    """
    Verifica se o arquivo é válido para processamento.
    Retorna (True, "") se OK, ou (False, mensagem_erro) se inválido.
    """
    path = Path(caminho)

    if not path.exists():
        return False, "O arquivo não foi encontrado no caminho informado."

    if not path.is_file():
        return False, "O caminho selecionado não é um arquivo."

    if path.suffix.lower() not in (".xlsx", ".csv"):
        return False, f"Formato '{path.suffix}' não suportado. Use .xlsx ou .csv."

    if path.stat().st_size == 0:
        return False, "O arquivo está vazio (0 bytes)."

    if path.stat().st_size > 50 * 1024 * 1024:  # 50 MB
        return False, "Arquivo muito grande (limite: 50 MB)."

    return True, ""


# ════════════════════════════════════════════════════════
#  GERADOR DE NOME DE SAÍDA
# ════════════════════════════════════════════════════════
def gerar_nome_saida(caminho_original: str, tipo: str) -> str:
    """
    Gera o caminho do arquivo de saída com nome automático.
    Formato: NomeOriginal_organizado_TIPO_YYYYMMDD_HHMM.xlsx
    Salva na mesma pasta do original.
    """
    path = Path(caminho_original)
    pasta = path.parent
    nome_base = path.stem

    # Remove caracteres especiais do nome base
    nome_base_limpo = re.sub(r"[^\w\-]", "_", nome_base)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    nome_saida = f"{nome_base_limpo}_organizado_{tipo}_{timestamp}.xlsx"

    return str(pasta / nome_saida)


# ════════════════════════════════════════════════════════
#  LIMPEZA GENÉRICA DE DATAFRAME
# ════════════════════════════════════════════════════════
def limpeza_basica(df):
    """
    Aplica limpeza básica em qualquer DataFrame:
    - Remove linhas completamente vazias
    - Remove colunas completamente vazias
    - Strip em colunas de texto
    - Remove duplicatas exatas
    Retorna (df_limpo, estatísticas_dict)
    """
    import pandas as pd

    stats = {
        "linhas_originais": len(df),
        "duplicados": 0,
        "vazias": 0,
        "colunas_criadas": [],
    }

    # Remove colunas totalmente vazias
    df = df.dropna(axis=1, how="all")

    # Remove linhas totalmente vazias
    antes = len(df)
    df = df.dropna(how="all")
    stats["vazias"] = antes - len(df)

    # Strip em valores de texto
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": None, "None": None, "": None})

    # Remove duplicatas exatas
    antes = len(df)
    df = df.drop_duplicates()
    stats["duplicados"] = antes - len(df)

    # Reseta índice
    df = df.reset_index(drop=True)

    stats["linhas_finais"] = len(df)
    return df, stats


# ════════════════════════════════════════════════════════
#  PADRONIZAÇÃO DE TEXTOS
# ════════════════════════════════════════════════════════
def titulo_proprio(texto) -> str:
    """Converte string para Title Case, tratando None/nan."""
    if not texto or str(texto).strip() in ("nan", "None", ""):
        return ""
    return str(texto).strip().title()


def uppercase_seguro(texto) -> str:
    """Converte para MAIÚSCULAS com segurança."""
    if not texto or str(texto).strip() in ("nan", "None", ""):
        return ""
    return str(texto).strip().upper()


def lowercase_seguro(texto) -> str:
    """Converte para minúsculas com segurança."""
    if not texto or str(texto).strip() in ("nan", "None", ""):
        return ""
    return str(texto).strip().lower()


# ════════════════════════════════════════════════════════
#  FORMATAÇÃO DE VALORES MONETÁRIOS
# ════════════════════════════════════════════════════════
def limpar_valor_monetario(valor) -> float | None:
    """
    Tenta converter string com formato monetário para float.
    Ex: "R$ 1.234,56" → 1234.56
        "1,234.56"    → 1234.56
        "1234.56"     → 1234.56
    """
    if valor is None:
        return None
    try:
        return float(valor)
    except (ValueError, TypeError):
        pass

    texto = str(valor).strip()
    # Remove símbolos monetários e espaços
    texto = re.sub(r"[R$€£¥\s]", "", texto)

    # Detecta formato brasileiro (1.234,56) vs americano (1,234.56)
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            # Formato BR: ponto como milhar, vírgula como decimal
            texto = texto.replace(".", "").replace(",", ".")
        else:
            # Formato US: vírgula como milhar, ponto como decimal
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except (ValueError, TypeError):
        return None


# ════════════════════════════════════════════════════════
#  PADRONIZAÇÃO DE DATAS
# ════════════════════════════════════════════════════════
def padronizar_data(valor, formato_saida: str = "%d/%m/%Y"):
    """
    Tenta interpretar diversos formatos de data e padroniza.
    Retorna string formatada ou None se não conseguir.
    """
    import pandas as pd

    if valor is None or str(valor).strip() in ("nan", "None", ""):
        return None
    try:
        dt = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.strftime(formato_saida)
    except Exception:
        return None


# ════════════════════════════════════════════════════════
#  PADRONIZAÇÃO DE CPF/CNPJ
# ════════════════════════════════════════════════════════
def formatar_cpf(valor) -> str:
    """Formata CPF: 12345678901 → 123.456.789-01"""
    if not valor:
        return ""
    nums = re.sub(r"\D", "", str(valor))
    if len(nums) == 11:
        return f"{nums[:3]}.{nums[3:6]}.{nums[6:9]}-{nums[9:]}"
    return str(valor)


def formatar_telefone(valor) -> str:
    """Formata telefone brasileiro (8 ou 9 dígitos + DDD)."""
    if not valor:
        return ""
    nums = re.sub(r"\D", "", str(valor))
    if len(nums) == 11:
        return f"({nums[:2]}) {nums[2:7]}-{nums[7:]}"
    if len(nums) == 10:
        return f"({nums[:2]}) {nums[2:6]}-{nums[6:]}"
    return str(valor)


# ════════════════════════════════════════════════════════
#  EXPORTAÇÃO COM ESTILO
# ════════════════════════════════════════════════════════
def exportar_xlsx_estilizado(df, caminho: str, nome_aba: str = "Dados"):
    """
    Exporta o DataFrame para Excel com formatação visual básica:
    - Cabeçalho azul escuro com texto branco e negrito
    - Linhas alternadas em cinza claro
    - Colunas com auto-largura
    - Bordas sutis
    """
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    import pandas as pd

    # Exporta com pandas primeiro
    df.to_excel(caminho, index=False, sheet_name=nome_aba)

    # Abre para estilizar
    wb = load_workbook(caminho)
    ws = wb[nome_aba]

    # ── Estilos ──────────────────────────────────────────
    fill_header   = PatternFill("solid", fgColor="1E3A5F")
    fill_linha_par = PatternFill("solid", fgColor="F0F4FA")
    font_header   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    font_normal   = Font(size=10, name="Calibri")
    alinha_centro = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alinha_esq    = Alignment(horizontal="left",   vertical="center")
    borda_fina    = Border(
        left=Side(style="thin", color="D0D7E3"),
        right=Side(style="thin", color="D0D7E3"),
        top=Side(style="thin", color="D0D7E3"),
        bottom=Side(style="thin", color="D0D7E3"),
    )

    # ── Cabeçalho ────────────────────────────────────────
    for cell in ws[1]:
        cell.fill   = fill_header
        cell.font   = font_header
        cell.alignment = alinha_centro
        cell.border = borda_fina

    # ── Linhas de dados ──────────────────────────────────
    for row_idx in range(2, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.font   = font_normal
            cell.border = borda_fina
            cell.alignment = alinha_esq
            if row_idx % 2 == 0:
                cell.fill = fill_linha_par

    # ── Auto-largura de colunas ──────────────────────────
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        col_letra = get_column_letter(col_idx)
        # Limita entre 10 e 45 caracteres
        ws.column_dimensions[col_letra].width = min(max(max_len + 4, 10), 45)

    # ── Altura padrão das linhas ─────────────────────────
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    # ── Congela o cabeçalho ──────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(caminho)
